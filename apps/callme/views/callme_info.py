import xlrd
import os

from django.contrib.auth import get_user, get_user_model

from django_filters import CharFilter
from django_filters import rest_framework as filters
from rest_framework import viewsets, permissions

from rest_framework import generics
from rest_framework.generics import get_object_or_404
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.views import APIView
from rest_framework.response import Response

from apps.callme.models import Company, PropertyInfo, OfferStatus, PropertyFileInfo
from apps.callme.serializers import (
    CallMeInfoSerializer,
    OfferStatusSerializer,
    PropertyFileSerializer,
    CommentOfferTabCustomerSerializer,
    CommentOfferTabClientSerializer,
    CommentOfferTabAgentSerializer,
)
from apps.authentication.models import Client
from apps.callme.models import (
    CommentOfferTabCustomer,
    CommentOfferTabClient,
    CommentOfferTabAgent,
)


from openpyxl import load_workbook

User = get_user_model()


__all__ = (
    "CallMeInfoViewSet",
    "OfferStatusViewSet",
    "PropertyInfoViewSet",
    "FileUploadView",
    "CreatePropertyInfoCustomerComment",
    "CreatePropertyInfoClientComment",
    "CreatePropertyInfoAgentComment",
)


class OfferStatusViewSet(viewsets.ModelViewSet):
    queryset = OfferStatus.objects.all()
    serializer_class = OfferStatusSerializer
    permission_classes = [permissions.IsAuthenticated]


class CallMeInfoFilter(filters.FilterSet):
    apn = CharFilter(field_name="apn", lookup_expr="icontains")
    reference = CharFilter(field_name="reference", lookup_expr="icontains")
    company_name = CharFilter(field_name="company_name", lookup_expr="icontains")

    class Meta:
        model = PropertyInfo
        fields = ("apn", "reference", "company_name")


class CallMeInfoViewSet(viewsets.ModelViewSet):
    serializer_class = CallMeInfoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = CallMeInfoFilter

    def get_queryset(self):
        current_user = self.request.user
        user = User.objects.filter(username=current_user).first()
        client = Client.objects.filter(user=user).first()
        print("User: ", client.client_code)
        if (
            current_user.designation_category == "current_client"
            or current_user.designation_category == "new_client"
            or current_user.designation_category == "affiliate_partner"
        ):
            qs = PropertyInfo.objects.filter(client_code=client.client_code)
            return qs
        elif current_user.is_superuser:
            qs = PropertyInfo.objects.all()
            return qs


class PropertyInfoViewSet(viewsets.ModelViewSet):
    serializer_class = CallMeInfoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        current_user = self.request.user
        user = User.objects.filter(username=current_user).first()
        client = Client.objects.filter(user=user).first()
        print("User: ", client.client_code)
        if (
            current_user.designation_category == "current_client"
            or current_user.designation_category == "new_client"
            or current_user.designation_category == "affiliate_partner"
        ):
            qs = PropertyInfo.objects.filter(client_code=client.client_code)
            return qs
        elif current_user.is_superuser:
            qs = PropertyInfo.objects.all()
            return qs


class CreatePropertyInfoCustomerComment(generics.CreateAPIView):
    serializer_class = CommentOfferTabCustomerSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CommentOfferTabCustomer.objects.select_related(
        "property_info", "user"
    ).all()

    def perform_create(self, serializer):
        user = self.request.user
        property_info_id = self.kwargs.get("id")
        property_info = get_object_or_404(PropertyInfo, id=property_info_id)
        serializer.save(user=user, property_info=property_info)


class CreatePropertyInfoClientComment(generics.CreateAPIView):
    serializer_class = CommentOfferTabClientSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CommentOfferTabClient.objects.select_related(
        "property_info", "user"
    ).all()

    def perform_create(self, serializer):
        user = self.request.user
        property_info_id = self.kwargs.get("id")
        property_info = get_object_or_404(PropertyInfo, id=property_info_id)
        serializer.save(user=user, property_info=property_info)


class CreatePropertyInfoAgentComment(generics.CreateAPIView):
    serializer_class = CommentOfferTabAgentSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CommentOfferTabAgent.objects.select_related(
        "property_info", "user"
    ).all()

    def perform_create(self, serializer):
        user = self.request.user
        property_info_id = self.kwargs.get("id")
        property_info = get_object_or_404(PropertyInfo, id=property_info_id)
        serializer.save(user=user, property_info=property_info)


class FileUploadView(APIView):
    parser_classes = (FormParser, MultiPartParser)
    serializer_class = PropertyFileSerializer

    def post(self, request, format=None):
        file_obj = request.data["file"]
        username = request.data["username"]
        user = User.objects.filter(username=username).first()
        client = Client.objects.filter(user=user).first()
        wb_obj = load_workbook(filename=file_obj)
        sheet_name = "Sheet1"
        sheet = wb_obj[sheet_name]

        for sheet_name in wb_obj.sheetnames:
            for value in sheet.iter_rows(values_only=True):
                full_name = value[0]
                company_name = value[1]
                reference_number = value[2]
                apn = value[3]
                county = value[4]
                state = value[5]
                size = value[6]
                address = value[7]
                price = value[8]
                due_diligence = value[9]
                ad_content = value[10]
                images = value[11]
                website = value[12]
                comment_offer_tab_customer = value[13]
                comment_offer_tab_client = value[14]
                comment_sales_agent_notes = value[15]
                facebook = value[16]
                fb_groups = value[17]
                landmodo = value[18]
                fsbo = value[19]
                instagram = value[20]
                land_listing = value[21]
                land_flip = value[22]
                land_hub = value[23]
                land_century = value[24]

                property_info = PropertyInfo.objects.create(
                    client_code=client.client_code,
                    full_name=full_name,
                    company_name=company_name,
                    reference_number=reference_number,
                    apn=apn,
                    county=county,
                    state=state,
                    size=size,
                    address=address,
                    price=price,
                    due_diligence=due_diligence,
                    ad_content=ad_content,
                    images=images,
                    website=website,
                    facebook=facebook,
                    fb_groups=fb_groups,
                    landmodo=landmodo,
                    fsbo=fsbo,
                    instagram=instagram,
                    land_listing=land_listing,
                    land_flip=land_flip,
                    land_hub=land_hub,
                    land_century=land_century,
                )
                property_info_id = PropertyInfo.objects.filter(id=property_info.id).first()
                CommentOfferTabCustomer.objects.create(
                    property_info=property_info_id,
                    comment=comment_offer_tab_customer,
                    user=user,
                )
                CommentOfferTabClient.objects.create(
                    property_info=property_info_id,
                    comment=comment_offer_tab_client,
                    user=user,
                )
                CommentOfferTabAgent.objects.create(
                    property_info=property_info_id,
                    comment=comment_sales_agent_notes,
                    user=user,
                )
        return Response(status=200)
